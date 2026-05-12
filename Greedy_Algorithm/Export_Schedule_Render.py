import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl.chart import BarChart, Reference

def Export_Schedule_Render(
    rows,
    event_state,
    dict_events,
    dict_employees,
    input_path,
    period_start=None,
    period_end=None
):

    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    DATA_DIR = os.path.join(BASE_DIR, "Data")
    os.makedirs(DATA_DIR, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(input_path))[0]
    output_path = os.path.join(DATA_DIR, f"{base_name}_vaktaplan.xlsx")

    # =========================
    # BUILD DATA
    # =========================
    schedule_rows = []

    for row in rows:
        event = dict_events[row["EventID"]]
        employee = dict_employees[row["EmployeeID"]]

        schedule_rows.append({
            "Event": event.get("Event", ""),
            "Hall": event.get("Hall", ""),
            "Date": pd.to_datetime(event["Date"]),
            "Start": str(event["ShiftBegins"]),
            "End": str(event["ShiftEnds"]),
            "Employee": employee.get("EmployeeName", "")
        })

    for event_id, event in dict_events.items():

        state = event_state.get(event_id, {})

        assigned = state.get("Assigned", 0)
        required = event.get("Employees", 0)
        missing = required - assigned

        for _ in range(max(0, missing)):
            schedule_rows.append({
                "Event": event.get("Event", ""),
                "Hall": event.get("Hall", ""),
                "Date": pd.to_datetime(event["Date"]),
                "Start": str(event["ShiftBegins"]),
                "End": str(event["ShiftEnds"]),
                "Employee": None   
            })

    df = pd.DataFrame(schedule_rows)

    if df.empty:
        return output_path

    df = df.sort_values(["Date", "Start"])

    grouped = (
        df.groupby(["Event", "Date", "Start", "End", "Hall"])["Employee"]
        .apply(lambda x: sorted([v for v in x if pd.notna(v)]))
        .reset_index()
    )

    grouped = grouped.sort_values(["Date", "Start"]).reset_index(drop=True)

    # =========================
    # EMPLOYEE VIEW DATA
    # =========================
    emp_grouped = (
        df[df["Employee"].notna()]
        .groupby("Employee")[["Event", "Date", "Start", "End", "Hall"]]
        .apply(lambda x: list(zip(x["Event"], x["Date"], x["Start"], x["End"], x["Hall"])))
    )

    emp_grouped = dict(sorted(emp_grouped.items()))

    unassigned = df[df["Employee"].isna()]

    if not unassigned.empty:
        emp_grouped["Ómannaðar vaktir"] = list(zip(
            unassigned["Event"],
            unassigned["Date"],
            unassigned["Start"],
            unassigned["End"],
            unassigned["Hall"]
        ))

    # =========================
    # STATS DATA
    # =========================
    shift_counts = df.groupby("Employee").size()

    def calc_hours(row):
        start = pd.to_datetime(row["Start"])
        end = pd.to_datetime(row["End"])
        if end < start:
            end += pd.Timedelta(days=1)
        return (end - start).total_seconds() / 3600

    df["Hours"] = df.apply(calc_hours, axis=1)
    hours_counts = df.groupby("Employee")["Hours"].sum()

    availability = {}
    for emp_id, emp in dict_employees.items():
        name = emp.get("EmployeeName", "")
        availability[name] = round(emp.get("Availability_ratio", 0), 2)

    # =========================
    # EXPORT
    # =========================
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        wb = writer.book
        ws = wb.create_sheet("Events")
        ws_emp = wb.create_sheet("Employees")
        ws_cal = wb.create_sheet("Calendar")
        ws_stats = wb.create_sheet("Stats")

        bold = Font(bold=True)
        fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        orange = PatternFill(start_color="FF6E1B", end_color="FF6E1B", fill_type="solid")
        border = Border(bottom=Side(style="thin"))
        center = Alignment(horizontal="center", vertical="center")

        months = [
            "janúar", "febrúar", "mars", "apríl", "maí", "júní",
            "júlí", "ágúst", "september", "október", "nóvember", "desember"
        ]

        # =========================
        # EVENTS SHEET
        # =========================
        col = 1
        for _, row in grouped.iterrows():

            ws.cell(row=1, column=col).value = f"{row['Event']} ({row['Hall']})"
            ws.cell(row=2, column=col).value = f"{row['Date'].day}. {months[row['Date'].month-1]} {row['Date'].year}"
            ws.cell(row=3, column=col).value = f"{row['Start']} - {row['End']}"

            for r in [1, 2, 3]:
                c = ws.cell(row=r, column=col)
                c.font = bold
                c.fill = fill
                c.alignment = center

            ws.cell(row=3, column=col).border = border

            for i, name in enumerate(row["Employee"]):
                if name is None:
                    continue
                ws.cell(row=4 + i, column=col).value = name

            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 25
            col += 1

        # =========================
        # EMPLOYEES SHEET
        # =========================
        col = 1
        for emp, events in emp_grouped.items():

            ws_emp.cell(row=1, column=col).value = emp

            row_ptr = 2
            for event, date, start, end, hall in events:
                ws_emp.cell(row=row_ptr, column=col).value = f"{event} ({hall})"
                ws_emp.cell(row=row_ptr + 1, column=col).value = f"{date.day}. {months[date.month-1]} {date.year}"
                ws_emp.cell(row=row_ptr + 2, column=col).value = f"{start} - {end}"
                row_ptr += 3

            col += 1

        # =========================
        # STATS SHEET
        # =========================
        employees_sorted = sorted([str(e) for e in df["Employee"].dropna().unique()])

        if df["Employee"].isna().any():
            employees_sorted = ["Ómannaðar vaktir"] + employees_sorted

        ws_stats["A1"] = "Starfsmaður"
        ws_stats["C1"] = "Fjöldi vakta"

        for i, emp in enumerate(employees_sorted, start=2):

            if emp == "Ómannaðar vaktir":
                ws_stats.cell(row=i, column=1).value = emp
                ws_stats.cell(row=i, column=3).value = int(shift_counts.get(None, 0))
                continue

            ws_stats.cell(row=i, column=1).value = emp
            ws_stats.cell(row=i, column=3).value = int(shift_counts.get(emp, 0))

    return output_path